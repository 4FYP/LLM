#!/usr/bin/env python3.12
import os, io, json, base64, uuid, csv
import httpx
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional
from dotenv import load_dotenv
from odoo_client import OdooClient

# Downloads folder (next to app.py)
DOWNLOADS_DIR = Path(__file__).parent / "downloads"
DOWNLOADS_DIR.mkdir(exist_ok=True)

load_dotenv()

DEEPSEEK_API_KEY  = os.getenv("DEEPSEEK_API_KEY")
DEEPSEEK_MODEL    = os.getenv("DEEPSEEK_MODEL", "deepseek-chat")
DEEPSEEK_API_BASE = os.getenv("DEEPSEEK_API_BASE", "https://api.deepseek.com")

ODOO_URL = os.getenv("ODOO_URL", "")
ODOO_DB  = os.getenv("ODOO_DB",  "")

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")

# session_id → {"client": OdooClient, "db": str, "username": str, "version": str}
sessions: dict[str, dict] = {}

# ── Odoo tool definitions for DeepSeek ───────────────────────────────────────
ODOO_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "odoo_search",
            "description": (
                "Fetch records from an Odoo model. Use for sales, invoices, "
                "customers, products, inventory, HR, projects, etc."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {
                        "type": "string",
                        "description": (
                            "Odoo model e.g. 'sale.order', 'account.move', "
                            "'res.partner', 'product.product', 'stock.picking', "
                            "'purchase.order', 'hr.employee', 'project.task'"
                        ),
                    },
                    "domain": {
                        "type": "array",
                        "description": "Filter domain e.g. [['state','=','sale'],['amount_total','>',1000]]",
                        "default": [],
                    },
                    "fields": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Fields to return e.g. ['name','partner_id','amount_total','state']",
                    },
                    "limit": {"type": "integer", "default": 50},
                    "order": {
                        "type": "string",
                        "description": "e.g. 'date_order desc'",
                        "default": "",
                    },
                },
                "required": ["model", "fields"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_count",
            "description": "Count records matching a domain in an Odoo model.",
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {"type": "string"},
                    "domain": {"type": "array", "default": []},
                },
                "required": ["model"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_fields",
            "description": "List available fields for an Odoo model so you know what to query.",
            "parameters": {
                "type": "object",
                "properties": {"model": {"type": "string"}},
                "required": ["model"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_create",
            "description": (
                "Create a new record in Odoo. Returns the new record ID.\n"
                "For One2many line fields (invoice_line_ids, order_line, etc.) use:\n"
                "  [[0, 0, {field: value, ...}]]  — one entry per line\n"
                "Examples:\n"
                "  Vendor bill: model='account.move', values={move_type:'in_invoice', partner_id:ID, "
                "invoice_date:'2026-06-13', currency_id:CUR_ID, "
                "invoice_line_ids:[[0,0,{name:'desc',price_unit:1000,quantity:1,account_id:ACC_ID}]]}\n"
                "  Sale order:  model='sale.order',   values={partner_id:ID, "
                "order_line:[[0,0,{product_id:PID,product_uom_qty:1,price_unit:500}]]}\n"
                "  Purchase:    model='purchase.order',values={partner_id:ID, "
                "order_line:[[0,0,{product_id:PID,product_qty:1,price_unit:200,name:'desc',date_planned:'2026-06-20'}]]}"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {"type": "string", "description": "Odoo model name"},
                    "values": {
                        "type": "object",
                        "description": "Field values dict. Use [[0,0,{...}]] for One2many fields.",
                    },
                },
                "required": ["model", "values"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_write",
            "description": (
                "Update existing Odoo records by their IDs.\n"
                "For One2many lines: [[1,line_id,{field:val}]] update, [[2,line_id]] delete, [[0,0,{...}]] add new.\n"
                "Examples:\n"
                "  Update partner email: model='res.partner', ids=[5], values={email:'new@email.com'}\n"
                "  Set inventory qty:    model='stock.quant',  ids=[12], values={inventory_quantity:50}"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {"type": "string"},
                    "ids":   {"type": "array", "items": {"type": "integer"}, "description": "Record IDs to update"},
                    "values":{"type": "object", "description": "Fields to update"},
                },
                "required": ["model", "ids", "values"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_unlink",
            "description": "Delete Odoo records permanently by their IDs. Use with caution.",
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {"type": "string"},
                    "ids":   {"type": "array", "items": {"type": "integer"}},
                },
                "required": ["model", "ids"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_action",
            "description": (
                "Call a method/action on Odoo records (workflow transitions).\n"
                "Common methods:\n"
                "  action_post              → Confirm/Post invoice or bill\n"
                "  button_confirm           → Confirm purchase order\n"
                "  action_confirm           → Confirm sale order\n"
                "  action_validate          → Validate payment\n"
                "  button_validate          → Validate delivery (stock.picking)\n"
                "  action_apply_inventory   → Apply inventory adjustment (stock.quant)\n"
                "  action_set_quantities_to_reservation → Set done qty in delivery\n"
                "  action_cancel            → Cancel a record\n"
                "  action_draft             → Reset to draft\n"
                "Example: Confirm invoice → model='account.move', method='action_post', ids=[42]"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model":  {"type": "string"},
                    "method": {"type": "string", "description": "Method name to call"},
                    "ids":    {"type": "array",  "items": {"type": "integer"}, "description": "Record IDs (can be [] for global actions)"},
                    "kwargs": {"type": "object",  "description": "Optional keyword args", "default": {}},
                },
                "required": ["model", "method", "ids"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_export",
            "description": (
                "Export Odoo records as a downloadable CSV file. "
                "Returns a download_url the user can click. "
                "Use this whenever the user asks to export, download, or get a file of data."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model":    {"type": "string", "description": "Odoo model name"},
                    "domain":   {"type": "array",  "description": "Filter domain", "default": []},
                    "fields":   {"type": "array",  "items": {"type": "string"}, "description": "Fields to export"},
                    "filename": {"type": "string", "description": "Output filename e.g. 'customer_invoices.csv'"},
                    "limit":    {"type": "integer","description": "Max records (default 5000)", "default": 5000},
                },
                "required": ["model", "fields"],
            },
        },
    },
]

SYSTEM_WITH_ODOO = """You are an expert Odoo ERP assistant with FULL access to a live Odoo database.
You can READ, CREATE, UPDATE, DELETE records and trigger workflow actions.

═══ TOOLS ═══
odoo_search   → Read/fetch records
odoo_count    → Count records
odoo_fields   → Discover model fields
odoo_create   → Create new record (returns ID)
odoo_write    → Update existing records
odoo_unlink   → Delete records
odoo_export   → Export data as downloadable CSV (returns download_url)
odoo_action   → Trigger workflow (confirm, validate, post, cancel…)

═══ CURRENCY RULE — CRITICAL ═══
- ALWAYS include 'currency_id' in fields for financial models
- currency_id returns [id, "AED"] — use currency_id[1] for display, NEVER assume "$"

═══ CREATE OPERATIONS ═══

Vendor Bill:
  1. odoo_search res.partner [['name','ilike','vendor']] ['id','name']
  2. odoo_search account.account [['account_type','=','expense']] ['id','name'] limit=5
  3. odoo_create account.move {move_type:'in_invoice', partner_id:ID, invoice_date:'YYYY-MM-DD',
       invoice_line_ids:[[0,0,{name:'desc', price_unit:1000, quantity:1, account_id:ACC_ID}]]}
  4. (optional) odoo_action account.move 'action_post' [new_id]  ← confirms the bill

Customer Invoice:
  Same as bill but move_type:'out_invoice'

Sale Order:
  1. odoo_search res.partner [['customer_rank','>',0]] ['id','name']
  2. odoo_search product.product [['name','ilike','product']] ['id','name','list_price']
  3. odoo_create sale.order {partner_id:ID, order_line:[[0,0,{product_id:PID, product_uom_qty:1, price_unit:PRICE}]]}
  4. odoo_action sale.order 'action_confirm' [new_id]

Purchase Order:
  1. Find vendor: odoo_search res.partner [['supplier_rank','>',0]]
  2. odoo_create purchase.order {partner_id:ID, order_line:[[0,0,{product_id:PID, product_qty:QTY, price_unit:PRICE, name:'desc', date_planned:'YYYY-MM-DD'}]]}
  3. odoo_action purchase.order 'button_confirm' [new_id]

═══ UPDATE OPERATIONS ═══
  Update any field: odoo_write model [id1,id2] {field: new_value}

═══ INVENTORY ADJUSTMENT ═══
  1. odoo_search product.product [['name','ilike','product']] ['id','name']
  2. odoo_search stock.quant [['product_id','=',PID],['location_id.usage','=','internal']] ['id','quantity','inventory_quantity']
  3. odoo_write stock.quant [quant_id] {inventory_quantity: NEW_QTY}
  4. odoo_action stock.quant 'action_apply_inventory' [quant_id]

═══ WORKFLOW ACTIONS ═══
  Confirm invoice:      odoo_action account.move   'action_post'          [id]
  Confirm sale:         odoo_action sale.order     'action_confirm'       [id]
  Confirm purchase:     odoo_action purchase.order 'button_confirm'       [id]
  Validate delivery:    odoo_action stock.picking  'button_validate'      [id]
  Validate payment:     odoo_action account.payment 'action_validate'     [id]
  Cancel:               odoo_action <model>        'action_cancel'        [id]
  Reset to draft:       odoo_action <model>        'action_draft'         [id]

═══ COMMON MODELS ═══
  account.move     — Invoices/Bills     (move_type: out_invoice/in_invoice/out_refund/in_refund)
  sale.order       — Sales orders
  purchase.order   — Purchase orders
  res.partner      — Customers/Vendors
  product.product  — Products
  product.template — Product templates
  stock.picking    — Deliveries/Receipts
  stock.quant      — Inventory quantities
  account.payment  — Payments
  account.account  — Chart of accounts
  account.journal  — Journals
  hr.employee      — Employees
  project.task     — Tasks
  mrp.production   — Manufacturing orders

═══ EXPORT / DOWNLOAD ═══
  When user asks to export, download, or get a file:
  1. Use odoo_export with model, domain, fields, and a descriptive filename (e.g. 'invoices_2024.csv')
  2. The tool returns {download_url: '/download/filename.csv', count: N}
  3. Present the link as: [Download invoices_2024.csv](/download/invoices_2024.csv)
  4. ALWAYS format the link in markdown link syntax so the user can click it directly.

Always confirm with user before deleting. Show created record ID after creation. Answer in user's language."""

SYSTEM_PLAIN = (
    "You are an expert Odoo ERP assistant. "
    "Help with modules, development, configuration, and business processes."
)

MIME_MAP = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png",
            "gif":"image/gif","webp":"image/webp"}
TEXT_EXTS = {"txt","md","csv","json","xml","py","js","ts",
             "html","css","yaml","yml","ini","log","env","sql"}


# ── Pydantic models ───────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str

class OdooConnectRequest(BaseModel):
    url: str
    db: str
    username: str
    password: str

class Message(BaseModel):
    role: str
    content: str

class Attachment(BaseModel):
    type: str
    content: str
    mime: Optional[str] = None
    filename: str

class ChatRequest(BaseModel):
    messages: list[Message]
    attachment: Optional[Attachment] = None
    session_id: Optional[str] = None


# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/login")

@app.get("/login")
async def login_page():
    return FileResponse("login.html")

@app.get("/config")
async def get_config():
    return {"odoo_url": ODOO_URL, "odoo_db": ODOO_DB}

@app.get("/chat")
async def chat_root():
    return FileResponse("index.html")

@app.get("/chat/{chat_id}")
async def chat_by_id(chat_id: str):
    return FileResponse("index.html")


@app.get("/download/{filename}")
async def download_file(filename: str):
    """Serve generated export files."""
    # Security: no path traversal
    safe = Path(filename).name
    filepath = DOWNLOADS_DIR / safe
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found")
    media = "text/csv" if safe.endswith(".csv") else "application/octet-stream"
    return FileResponse(str(filepath), filename=safe, media_type=media)


@app.post("/login")
async def platform_login(req: LoginRequest):
    users_file = Path(__file__).parent / "users.json"
    try:
        users = json.loads(users_file.read_text())
    except Exception:
        users = [{"username": "admin", "password": "admin123", "name": "Admin"}]
    user = next((u for u in users if u["username"] == req.username and u["password"] == req.password), None)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password.")
    token = str(uuid.uuid4())
    return {"token": token, "username": user["username"], "name": user.get("name", user["username"])}


@app.post("/odoo/connect")
async def odoo_connect(req: OdooConnectRequest):
    try:
        client = OdooClient(req.url, req.db, req.username, req.password)
        sid = str(uuid.uuid4())
        sessions[sid] = {
            "client": client,
            "db": req.db,
            "url": req.url,
            "username": req.username,
            "version": client.server_version,
        }
        return {
            "session_id": sid,
            "db": req.db,
            "url": req.url,
            "username": req.username,
            "server_version": client.server_version,
        }
    except PermissionError as e:
        raise HTTPException(status_code=401, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Connection failed: {e}")


@app.get("/session/{session_id}")
async def check_session(session_id: str):
    if session_id in sessions:
        s = sessions[session_id]
        return {"valid": True, "db": s["db"], "url": s["url"], "username": s["username"]}
    return {"valid": False}

@app.post("/logout")
async def logout(body: dict):
    sid = body.get("session_id", "")
    sessions.pop(sid, None)
    return {"ok": True}


@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    data = await file.read()
    filename = file.filename or "file"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in MIME_MAP:
        return {"type":"image","content":base64.b64encode(data).decode(),
                "mime":MIME_MAP[ext],"filename":filename}

    if ext in ("xlsx","xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            parts = []
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = [r for r in ws.iter_rows(values_only=True)
                        if any(c is not None for c in r)]
                if not rows: continue
                parts.append(f"=== Sheet: {sname} ===")
                ncols = len(rows[0])
                str_rows = [[str(c) if c is not None else "" for c in r] for r in rows]
                widths = [max(len(r[i]) for r in str_rows) for i in range(ncols)]
                for r in str_rows:
                    parts.append(" | ".join(c.ljust(widths[i]) for i, c in enumerate(r)))

                # Python-computed column totals (accurate, AI must not recalculate)
                headers = [str(rows[0][i]) if rows[0][i] is not None else f"Col{i}" for i in range(ncols)]
                totals = {}
                for ci in range(ncols):
                    s = sum(r[ci] for r in rows[1:] if isinstance(r[ci], (int, float)) and not isinstance(r[ci], bool))
                    if s:
                        totals[headers[ci]] = int(s) if s == int(s) else round(s, 4)
                if totals:
                    parts.append("Totals: " + ", ".join(f"{k}={v}" for k, v in totals.items()))

            text = "\n".join(parts) or "[Empty Excel file]"
        except Exception as e:
            text = f"[Excel parse error: {e}]"
        return {"type":"text","content":text,"filename":filename}

    if ext == "pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            text = "\n\n".join(p.extract_text() or "" for p in reader.pages).strip()
            if not text: text = "[PDF has no extractable text]"
        except Exception as e:
            text = f"[PDF error: {e}]"
        return {"type":"text","content":text,"filename":filename}

    if ext == "docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs if p.text)
        except Exception as e:
            text = f"[DOCX error: {e}]"
        return {"type":"text","content":text,"filename":filename}

    if ext in TEXT_EXTS:
        return {"type":"text","content":data.decode("utf-8",errors="replace"),"filename":filename}

    try:
        return {"type":"text","content":data.decode("utf-8",errors="replace"),"filename":filename}
    except Exception:
        return {"type":"unsupported","content":"","filename":filename}


# ── DeepSeek helpers ──────────────────────────────────────────────────────────
async def deepseek_tool_call(messages: list, tools: list) -> dict:
    """Non-streaming call — used for tool-calling loop."""
    url = f"{DEEPSEEK_API_BASE}/chat/completions"
    headers = {"Authorization": f"Bearer {DEEPSEEK_API_KEY}",
               "Content-Type": "application/json"}
    payload = {
        "model": DEEPSEEK_MODEL,
        "messages": messages,
        "tools": tools,
        "tool_choice": "auto",
        "stream": False,
        "temperature": 0.3,
    }
    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(url, json=payload, headers=headers)
        resp.raise_for_status()
        body = resp.text
        try:
            return json.loads(body)
        except json.JSONDecodeError as e:
            raise ValueError(f"DeepSeek returned invalid JSON: {e}\nBody: {body[:300]}")


def run_odoo_tool(name: str, args: dict, client: OdooClient):
    """Execute one tool call against Odoo."""
    if name == "odoo_search":
        return client.search_read(
            args["model"],
            args.get("domain", []),
            args.get("fields", []),
            args.get("limit", 50),
            args.get("order", ""),
        )
    if name == "odoo_count":
        return {"count": client.count(args["model"], args.get("domain", []))}
    if name == "odoo_fields":
        return client.fields_get(args["model"])

    if name == "odoo_create":
        new_id = client.create(args["model"], args["values"])
        # Fetch actual display name/reference from Odoo after creation
        try:
            records = client.search_read(args["model"], [["id","=",new_id]], ["name","display_name"], 1)
            ref = records[0].get("name") or records[0].get("display_name") if records else None
        except Exception:
            ref = None
        return {
            "success": True,
            "id": new_id,
            "reference": ref,
            "message": f"Created {args['model']} — DB ID: {new_id}" + (f", Reference: {ref}" if ref else ""),
        }

    if name == "odoo_write":
        ok = client.write(args["model"], args["ids"], args["values"])
        return {"success": ok,
                "message": f"Updated {len(args['ids'])} record(s) in {args['model']}"}

    if name == "odoo_unlink":
        ok = client.unlink(args["model"], args["ids"])
        return {"success": ok,
                "message": f"Deleted {len(args['ids'])} record(s) from {args['model']}"}

    if name == "odoo_action":
        result = client.call_method(
            args["model"], args["method"],
            args.get("ids", []), args.get("kwargs", {})
        )
        return {"success": True, "result": result,
                "message": f"Called {args['method']} on {args['model']} IDs {args.get('ids',[])}"}

    if name == "odoo_export":
        model    = args["model"]
        domain   = args.get("domain", [])
        fields   = args.get("fields", [])
        limit    = args.get("limit", 5000)
        fname    = args.get("filename") or f"{model.replace('.','_')}_export.csv"
        fname    = Path(fname).name  # no path traversal

        records = client.search_read(model, domain, fields, limit)

        filepath = DOWNLOADS_DIR / fname
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(fields)
            for rec in records:
                row = []
                for field in fields:
                    val = rec.get(field, "")
                    if isinstance(val, (list, tuple)):
                        val = val[1] if len(val) > 1 else ""
                    elif val is None or val is False:
                        val = ""
                    row.append(val)
                writer.writerow(row)

        url = f"/download/{fname}"
        return {
            "success": True,
            "download_url": url,
            "filename": fname,
            "records_count": len(records),
            "message": f"Export ready — {len(records)} records",
        }

    return {"error": f"Unknown tool: {name}"}


@app.post("/chat")
async def chat(req: ChatRequest):
    session = sessions.get(req.session_id or "")
    # If session_id was provided but not found — session expired (server restarted)
    if req.session_id and not session:
        raise HTTPException(status_code=401, detail="session_expired")
    odoo: Optional[OdooClient] = session["client"] if session else None
    system = SYSTEM_WITH_ODOO if odoo else SYSTEM_PLAIN

    # Build base messages
    base: list[dict] = [{"role": "system", "content": system}]
    for i, m in enumerate(req.messages):
        is_last = (i == len(req.messages) - 1 and m.role == "user")
        if is_last and req.attachment:
            att = req.attachment
            if att.type == "image":
                base.append({"role": "user", "content": [
                    {"type": "text", "text": m.content or "Describe this."},
                    {"type": "image_url", "image_url": {"url": f"data:{att.mime};base64,{att.content}"}},
                ]})
            else:
                full = f"[File: {att.filename}]\n\n{att.content}"
                if m.content:
                    full += f"\n\n---\n{m.content}"
                base.append({"role": "user", "content": full})
        else:
            base.append({"role": m.role, "content": m.content})

    async def event_stream():
        def status(msg: str) -> str:
            return json.dumps({"t": "s", "m": msg}) + "\n"

        def chunk(text: str) -> str:
            return json.dumps({"t": "c", "v": text}) + "\n"

        def error(msg: str) -> str:
            return json.dumps({"t": "c", "v": f"\n\n⚠️ Error: {msg}"}) + "\n"

        ACTION_ICONS = {
            "odoo_search": "🔍", "odoo_count": "🔢", "odoo_fields": "📋",
            "odoo_create": "➕", "odoo_write": "✏️", "odoo_unlink": "🗑️",
            "odoo_action": "⚡",
        }

        tool_msgs = list(base)

        try:
            # ── Tool-calling loop (max 10 iterations) ─────────────────────
            if odoo:
                for iteration in range(10):
                    try:
                        result = await deepseek_tool_call(tool_msgs, ODOO_TOOLS)
                    except Exception as e:
                        yield error(f"DeepSeek call failed: {e}")
                        yield json.dumps({"t": "e"}) + "\n"
                        return

                    choice = result["choices"][0]

                    # No more tool calls — yield content and finish
                    if choice["finish_reason"] != "tool_calls":
                        content = (choice["message"].get("content") or "").strip()
                        if content:
                            for i in range(0, len(content), 40):
                                yield chunk(content[i:i+40])
                        yield json.dumps({"t": "e"}) + "\n"
                        return

                    # Process tool calls
                    tool_msgs.append(choice["message"])

                    for tc in choice["message"].get("tool_calls", []):
                        fn = tc["function"]["name"]
                        raw_args = tc["function"].get("arguments", "{}")

                        # ── Safe JSON parse of arguments ───────────────────
                        try:
                            args = json.loads(raw_args)
                        except json.JSONDecodeError:
                            # Auto-fix common DeepSeek JSON mistakes
                            fixed = raw_args
                            # Python booleans/null → JSON
                            import re as _re
                            fixed = _re.sub(r'\bTrue\b',  'true',  fixed)
                            fixed = _re.sub(r'\bFalse\b', 'false', fixed)
                            fixed = _re.sub(r'\bNone\b',  'null',  fixed)
                            # Single quotes → double quotes (simple cases)
                            fixed = fixed.replace("'", '"')
                            # Trailing commas before ] or }
                            fixed = _re.sub(r',\s*([}\]])', r'\1', fixed)
                            try:
                                args = json.loads(fixed)
                            except json.JSONDecodeError as je2:
                                err_msg = f"Malformed tool args for {fn}: {je2}"
                                yield status(f"⚠️ {err_msg}")
                                tool_msgs.append({
                                    "role": "tool",
                                    "tool_call_id": tc["id"],
                                    "content": json.dumps({"error": err_msg}),
                                })
                                continue

                        model_name = args.get("model", fn)
                        icon = ACTION_ICONS.get(fn, "🔧")

                        if fn == "odoo_action":
                            verb = f"Calling {args.get('method', '?')} on"
                        elif fn == "odoo_create":
                            verb = "Creating in"
                        elif fn == "odoo_write":
                            verb = "Updating"
                        elif fn == "odoo_unlink":
                            verb = "Deleting from"
                        elif fn == "odoo_count":
                            verb = "Counting"
                        elif fn == "odoo_fields":
                            verb = "Inspecting"
                        else:
                            verb = "Querying"

                        yield status(f"{icon} {verb} {model_name}…")

                        # ── Execute tool safely ────────────────────────────
                        try:
                            data = run_odoo_tool(fn, args, odoo)
                            if isinstance(data, list):
                                yield status(f"✅ {model_name}: {len(data)} record(s)")
                            elif isinstance(data, dict) and "id" in data:
                                yield status(f"✅ Created {model_name} — ID: {data['id']}")
                            elif isinstance(data, dict) and data.get("success"):
                                yield status(f"✅ {data.get('message', model_name + ' done')}")
                            else:
                                yield status(f"✅ {model_name} done")
                        except Exception as e:
                            short_err = str(e)[:200]
                            data = {"error": str(e)}
                            yield status(f"⚠️ {model_name}: {short_err}")

                        tool_msgs.append({
                            "role": "tool",
                            "tool_call_id": tc["id"],
                            "content": json.dumps(data, default=str),
                        })

                yield status("⚠️ Reached tool call limit — summarising…")

            # ── Stream final response ──────────────────────────────────────
            url = f"{DEEPSEEK_API_BASE}/chat/completions"
            headers = {
                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": DEEPSEEK_MODEL,
                "messages": tool_msgs if odoo else base,
                "temperature": 0.7,
                "stream": True,
            }
            async with httpx.AsyncClient(timeout=120) as client:
                async with client.stream("POST", url, json=payload, headers=headers) as resp:
                    async for line in resp.aiter_lines():
                        if not line.startswith("data: "):
                            continue
                        raw = line[6:]
                        if raw == "[DONE]":
                            break
                        try:
                            delta = json.loads(raw)["choices"][0]["delta"].get("content", "")
                            if delta:
                                yield chunk(delta)
                        except Exception:
                            pass

        except Exception as top_err:
            # Catch-all: send readable error to frontend instead of dropping connection
            yield error(str(top_err))

        yield json.dumps({"t": "e"}) + "\n"

    return StreamingResponse(event_stream(), media_type="text/plain")
